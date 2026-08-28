#!/usr/bin/env python3
"""
OSDS-NSTP Form 2-B Ready-to-Print Generator using openpyxl
Generates/populates official CHED OSDS-NSTP Form 2-B (Detailed NSTP Masterlist of Enrollees and Graduates).
"""

import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

def find_logo_file(filenames):
    """Helper to locate logo image files across standard project paths."""
    search_dirs = ['.', 'public', '../public', 'backend/assets', '../backend/assets']
    for d in search_dirs:
        for f in filenames:
            full_path = os.path.join(d, f)
            if os.path.exists(full_path):
                return full_path
    return None


def create_base_form_b_template(template_path="OSDS-NSTP-Form-2-B_2.xlsx"):
    """Creates the official CHED OSDS-NSTP Form 2-B base template if not present."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CHED NSTP Form 2-B"

    # Enforce print setup
    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_setup.paperSize = 9  # A4 Size
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False  # Allows multiple pages vertically for long rosters
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

    # Styling definitions
    font_header_bold = Font(name="Arial", size=10, bold=True)
    font_title = Font(name="Arial", size=11, bold=True)
    font_col_header = Font(name="Arial", size=9, bold=True)
    font_sub_header = Font(name="Arial", size=8, bold=True)
    fill_header = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
    fill_sub = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Institutional Top Headers (Rows 1 to 3)
    ws['A1'] = "Republic of the Philippines"
    ws['A2'] = "Office of the President"
    ws['A3'] = "Commission on Higher Education"
    for r in range(1, 4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        ws[f'A{r}'].alignment = align_center
        ws[f'A{r}'].font = font_header_bold

    # 2. Form Title & Institutional Information
    ws['A5'] = "NSTP 1 / NSTP 2 ENROLLMENT LIST"
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=15)
    ws['A5'].alignment = align_center
    ws['A5'].font = font_title

    ws['A6'] = "Academic Year: 2024-2025"
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=15)
    ws['A6'].alignment = align_center
    ws['A6'].font = font_header_bold

    ws['A8'] = "Name of HEI: Cavite State University - Naic"
    ws['A8'].font = font_header_bold
    ws['M8'] = "Region: 4A - CALABARZON"
    ws['M8'].font = font_header_bold

    ws['A9'] = "Address: Bucana Malaki, Naic, Cavite"
    ws['A9'].font = font_header_bold
    ws['M9'] = "NSTP Components: CWTS / ROTC / LTS"
    ws['M9'].font = font_header_bold

    # 3. Column Header Matrix (Rows 11 and 12)
    headers_row11 = [
        (1, "No."),
        (2, "Student No."),
        (3, "Student Name"),
        (6, "Program"),
        (7, "Sex"),
        (8, "Birthdate"),
        (9, "Address"),
        (14, "Contact Number"),
        (15, "Email Address")
    ]

    for col_idx, text in headers_row11:
        ws.cell(row=11, column=col_idx, value=text)

    # Merges for Row 11 & 12 Headers
    ws.merge_cells('A11:A12')
    ws.merge_cells('B11:B12')
    ws.merge_cells('C11:E11') # Student Name (Surname, First Name, Middle Name)
    ws.cell(row=12, column=3, value="Surname")
    ws.cell(row=12, column=4, value="First Name")
    ws.cell(row=12, column=5, value="Middle Name")

    ws.merge_cells('F11:F12') # Program
    ws.merge_cells('G11:G12') # Sex
    ws.merge_cells('H11:H12') # Birthdate

    ws.merge_cells('I11:M11') # Address
    ws.merge_cells('I12:J12')
    ws.cell(row=12, column=9, value="Street/Barangay")
    ws.merge_cells('K12:L12')
    ws.cell(row=12, column=11, value="Municipality/City")
    ws.cell(row=12, column=13, value="Province")

    ws.merge_cells('N11:N12') # Contact Number
    ws.merge_cells('O11:O12') # Email Address

    # Apply styling & borders to Header Matrix (Rows 11 and 12)
    for r in [11, 12]:
        ws.row_dimensions[r].height = 22
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = align_center
            cell.font = font_col_header if r == 11 else font_sub_header
            cell.fill = fill_header if r == 11 else fill_sub

    # Set Column Widths for Perfect Layout
    col_widths = {
        'A': 6,   # No.
        'B': 16,  # Student No.
        'C': 18,  # Surname
        'D': 18,  # First Name
        'E': 16,  # Middle Name
        'F': 14,  # Program
        'G': 8,   # Sex
        'H': 13,  # Birthdate
        'I': 14,  # Street/Brgy
        'J': 10,
        'K': 12,  # Municipality/City
        'L': 10,
        'M': 14,  # Province
        'N': 16,  # Contact No.
        'O': 28   # Email Address
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(template_path)
    print(f"Template created: {template_path}")
    return template_path


def generate_ready_to_print_form_b(
    student_records, 
    template_path="OSDS-NSTP-Form-2-B_2.xlsx", 
    output_filename="Ready_To_Print_Form_2B.xlsx"
):
    """
    Naglo-load ng Form 2-B template, nag-i-inject ng logos, at nagsusulat ng student data.
    """
    # 1. I-load ang exact file verbatim (o lumikha kung wala pa)
    if not os.path.exists(template_path):
        create_base_form_b_template(template_path)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Enforce Landscape A4 Print Settings
    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_setup.paperSize = 9  # A4 Size
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

    # 2. I-inject ang Logos sa Header
    try:
        ched_logo_path = find_logo_file(['ched-logo.png', 'ched_logo.png', 'ched.png'])
        cvsu_logo_path = find_logo_file(['cvsu.png', 'cvsu_logo.png', 'cvsu-logo.png'])

        if ched_logo_path:
            ched_logo = Image(ched_logo_path)
            ched_logo.width, ched_logo.height = 70, 70
            ws.add_image(ched_logo, 'B1') # Column B para sa kaliwa

        if cvsu_logo_path:
            cvsu_logo = Image(cvsu_logo_path)
            cvsu_logo.width, cvsu_logo.height = 70, 70
            ws.add_image(cvsu_logo, 'N1') # Column N para sa kanan
    except Exception as err:
        print(f"⚠️ Notice: Logo loading notice: {err}")

    # 3. Setup ng formatting (Borders at Alignment)
    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_font = Font(name="Arial", size=9)

    # Base sa standard Form 2-B, ang student records ay nag-uumpisa sa Row 13
    start_row = 13 

    # 4. I-loop ang mga estudyante galing sa database
    for index, student in enumerate(student_records):
        current_row = start_row + index
        ws.row_dimensions[current_row].height = 20
        
        # Column A (1): Number / Bilang
        ws.cell(row=current_row, column=1, value=index + 1)
        
        # Column B (2): Student Number
        ws.cell(row=current_row, column=2, value=student.get('student_no', student.get('studentId', '')))
        
        # Columns C to E (3-5): Name
        ws.cell(row=current_row, column=3, value=student.get('surname', student.get('lastName', '')))
        ws.cell(row=current_row, column=4, value=student.get('first_name', student.get('firstName', '')))
        ws.cell(row=current_row, column=5, value=student.get('middle_name', student.get('middleName', '')))
        
        # Column F (6): Program / Course
        ws.cell(row=current_row, column=6, value=student.get('program', student.get('course', 'BSIT')))
        
        # Column G (7): Sex
        ws.cell(row=current_row, column=7, value=student.get('sex', student.get('gender', 'M')))
        
        # Column H (8): Birthdate
        ws.cell(row=current_row, column=8, value=student.get('birthdate', student.get('birthDate', '')))
        
        # Column I (9): Street/Barangay
        ws.cell(row=current_row, column=9, value=student.get('street_brgy', student.get('street', '')))
        
        # Column K (11): Municipality/City (Kadalasan naka-merge sa template ang I&J kaya sa K ang City)
        ws.cell(row=current_row, column=11, value=student.get('city', student.get('municipality', '')))
        
        # Column M (13): Province (Kadalasan naka-merge ang K&L kaya sa M ang Province)
        ws.cell(row=current_row, column=13, value=student.get('province', 'Cavite'))
        
        # Column N (14): Contact Number
        ws.cell(row=current_row, column=14, value=student.get('contact_no', student.get('contactNumber', '')))
        
        # Column O (15): Email Address
        ws.cell(row=current_row, column=15, value=student.get('email', ''))

        # 5. I-apply ang Borders para hindi mukhang floating ang text
        # (Lalagyan ng border mula Column A hanggang Column O)
        for col in range(1, 16): 
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.font = data_font
            
            # Left aligned ang pangalan at address para madaling basahin, the rest center
            if col in [3, 4, 5, 9, 11, 13]: 
                cell.alignment = left_aligned
            else:
                cell.alignment = center_aligned

    # 6. Signatories Footer
    sig_row = start_row + len(student_records) + 2
    ws.cell(row=sig_row, column=1, value="Prepared by: NSTP Department Coordinator").font = Font(name="Arial", size=9, bold=True)
    ws.cell(row=sig_row, column=6, value="Certified Correct: Campus NSTP Director").font = Font(name="Arial", size=9, bold=True)
    ws.cell(row=sig_row, column=12, value="Approved: Campus Administrator").font = Font(name="Arial", size=9, bold=True)

    # 7. I-save as ready-to-print file
    wb.save(output_filename)
    print(f"✅ Success! Ang print-ready Excel para sa Form 2-B ay nai-save bilang: {output_filename}")
    return output_filename


# ==========================================
# EXAMPLE USAGE & TEST
# ==========================================
if __name__ == "__main__":
    sample_students = [
        {
            'student_no': '2025-0001',
            'surname': 'Dela Cruz',
            'first_name': 'Juan',
            'middle_name': 'Santos',
            'program': 'BSIT',
            'sex': 'M',
            'birthdate': '2005-08-17',
            'street_brgy': 'Bancaan',
            'city': 'Naic',
            'province': 'Cavite',
            'contact_no': '09123456789',
            'email': 'juan@cvsu.edu.ph'
        },
        {
            'student_no': '2025-0002',
            'surname': 'Belen',
            'first_name': 'Richard',
            'middle_name': 'Mariño',
            'program': 'BSIT',
            'sex': 'M',
            'birthdate': '2005-08-17',
            'street_brgy': 'Halang',
            'city': 'Naic',
            'province': 'Cavite',
            'contact_no': '09987654321',
            'email': 'richard@cvsu.edu.ph'
        }
    ]

    generate_ready_to_print_form_b(sample_students)
