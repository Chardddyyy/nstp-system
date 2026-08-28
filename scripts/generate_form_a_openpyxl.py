#!/usr/bin/env python3
"""
OSDS-NSTP Form 2-A Ready-to-Print Generator using openpyxl
Generates/populates official CHED OSDS-NSTP Form 2-A (Summary Number of Enrollment and Graduates of NSTP).
"""

import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

def find_logo_file(filenames):
    """Helper to locate logo image files across standard project paths."""
    search_dirs = ['.', 'public', '../public', 'backend/assets', '../backend/assets']
    for d in search_dirs:
        for f in filenames:
            full_path = os.path.join(d, f)
            if os.path.exists(full_path):
                return full_path
    return None

def create_base_form_a_template(template_path='OSDS-NSTP-Form-2-A.xlsx'):
    """Creates the official CHED OSDS-NSTP Form 2-A base template if not present."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OSDS-NSTP Form 2-A"

    # Enforce print setup
    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_setup.paperSize = 9  # A4 Size
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

    # Insert Official Logos in Header
    cvsu_logo = find_logo_file(['cvsu.png', 'cvsu-logo.png'])
    ched_logo = find_logo_file(['ched-logo.png', 'ched.png'])

    if cvsu_logo:
        try:
            img_cvsu = Image(cvsu_logo)
            img_cvsu.width = 62
            img_cvsu.height = 62
            ws.add_image(img_cvsu, 'A1') # Left Header (CvSU Logo)
        except Exception as e:
            print(f"Notice: CvSU logo load ({e})")

    if ched_logo:
        try:
            img_ched = Image(ched_logo)
            img_ched.width = 62
            img_ched.height = 62
            ws.add_image(img_ched, 'Y1') # Right Header (CHED Logo)
        except Exception as e:
            print(f"Notice: CHED logo load ({e})")

    # Styling definitions
    font_header_bold = Font(name="Arial", size=10, bold=True)
    font_title = Font(name="Arial", size=11, bold=True)
    font_col_header = Font(name="Arial", size=9, bold=True)
    font_sub_header = Font(name="Arial", size=8, bold=True)
    fill_header = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft institutional green
    fill_sub = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Institutional Top Headers
    ws['A1'] = "Republic of the Philippines"
    ws['A2'] = "Office of the President"
    ws['A3'] = "Commission on Higher Education"
    for r in range(1, 4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=26)
        ws[f'A{r}'].alignment = align_center
        ws[f'A{r}'].font = font_header_bold

    # 2. Form Title
    ws['A5'] = "SUMMARY NUMBER OF ENROLLMENT AND GRADUATES OF NSTP"
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=26)
    ws['A5'].alignment = align_center
    ws['A5'].font = font_title

    ws['A6'] = "Academic Year: 2024-2025"
    ws['A6'].font = font_header_bold
    ws['W6'] = "Region: 4A - CALABARZON"
    ws['W6'].font = font_header_bold

    # 3. Column Header Matrix (Rows 8 to 11)
    # Row 8
    ws.cell(row=8, column=1, value="NAME OF HEI/CAMPUS")
    ws.cell(row=8, column=2, value="Classification\n(Private/Public)")
    ws.cell(row=8, column=3, value="ENROLLMENT")
    ws.cell(row=8, column=21, value="GRADUATES")

    ws.merge_cells(start_row=8, start_column=1, end_row=11, end_column=1) # Col A
    ws.merge_cells(start_row=8, start_column=2, end_row=11, end_column=2) # Col B
    ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=20) # Enrollment (C to T)
    ws.merge_cells(start_row=8, start_column=21, end_row=8, end_column=26) # Graduates (U to Z)

    # Row 9: Semesters & Summer
    ws.cell(row=9, column=3, value="1st Sem.")
    ws.merge_cells(start_row=9, start_column=3, end_row=9, end_column=8) # C-H

    ws.cell(row=9, column=9, value="2nd Sem.")
    ws.merge_cells(start_row=9, start_column=9, end_row=9, end_column=14) # I-N

    ws.cell(row=9, column=15, value="Summer")
    ws.merge_cells(start_row=9, start_column=15, end_row=9, end_column=20) # O-T

    ws.cell(row=9, column=21, value="")
    ws.merge_cells(start_row=9, start_column=21, end_row=9, end_column=26) # U-Z

    # Row 10: Tracks (ROTC, CWTS, LTS)
    for sem_start in [3, 9, 15, 21]:
        ws.cell(row=10, column=sem_start, value="ROTC")
        ws.merge_cells(start_row=10, start_column=sem_start, end_row=10, end_column=sem_start+1)
        ws.cell(row=10, column=sem_start+2, value="CWTS")
        ws.merge_cells(start_row=10, start_column=sem_start+2, end_row=10, end_column=sem_start+3)
        ws.cell(row=10, column=sem_start+4, value="LTS")
        ws.merge_cells(start_row=10, start_column=sem_start+4, end_row=10, end_column=sem_start+5)

    # Row 11: Gender (M, F)
    for col in range(3, 27, 2):
        ws.cell(row=11, column=col, value="M")
        ws.cell(row=11, column=col+1, value="F")

    # Apply styling & borders to Header Matrix (Rows 8 to 11, Cols 1 to 26)
    for r in range(8, 12):
        ws.row_dimensions[r].height = 20
        for c in range(1, 27):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = align_center
            cell.font = font_col_header if r <= 10 else font_sub_header
            cell.fill = fill_header if r == 8 else fill_sub

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 20
    for col_idx in range(3, 27):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 8.5

    # Default institutional row in Row 14
    ws['A14'] = "CVSU NAIC"
    ws['B14'] = "PUBLIC"

    wb.save(template_path)
    print(f"Template created: {template_path}")
    return template_path


def generate_ready_to_print_form_a(database_records, template_path='OSDS-NSTP-Form-2-A.xlsx'):
    """
    Populates official OSDS-NSTP-Form-2-A template with database records.
    """
    # 1. I-load ang mismong template file (lumikha kung wala pa)
    if not os.path.exists(template_path):
        create_base_form_a_template(template_path)

    wb = openpyxl.load_workbook(template_path)
    
    # Piliin ang active sheet
    ws = wb.active 

    # Enforce Landscape A4 Print Setup (1 Page Fit)
    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    data_font = Font(name="Arial", size=9)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # 2. I-set kung saang row mag-uumpisa ang data (Row 15 pababa base sa layout)
    # Ang Row 14 kasi sa template ay may nakalagay nang "CVSU NAIC | PUBLIC"
    start_row = 15 
    
    # 3. I-loop ang mga records galing sa database at ilagay sa Excel
    for index, record in enumerate(database_records):
        current_row = start_row + index
        ws.row_dimensions[current_row].height = 20
        
        # Isulat ang mga values sa kani-kanilang columns (A=1, B=2, C=3, atbp.)
        ws.cell(row=current_row, column=1, value=record.get('hei_name', ''))         # Column A
        ws.cell(row=current_row, column=2, value=record.get('classification', ''))   # Column B
        
        # 1ST SEMESTER (Columns C to H)
        ws.cell(row=current_row, column=3, value=record.get('sem1_rotc_m', 0))      # Column C (ROTC M)
        ws.cell(row=current_row, column=4, value=record.get('sem1_rotc_f', 0))      # Column D (ROTC F)
        ws.cell(row=current_row, column=5, value=record.get('sem1_cwts_m', 0))      # Column E (CWTS M)
        ws.cell(row=current_row, column=6, value=record.get('sem1_cwts_f', 0))      # Column F (CWTS F)
        ws.cell(row=current_row, column=7, value=record.get('sem1_lts_m', 0))       # Column G (LTS M)
        ws.cell(row=current_row, column=8, value=record.get('sem1_lts_f', 0))       # Column H (LTS F)
        
        # 2ND SEMESTER (Columns I to N)
        ws.cell(row=current_row, column=9, value=record.get('sem2_rotc_m', 0))      # Column I (ROTC M)
        ws.cell(row=current_row, column=10, value=record.get('sem2_rotc_f', 0))     # Column J (ROTC F)
        ws.cell(row=current_row, column=11, value=record.get('sem2_cwts_m', 0))     # Column K (CWTS M)
        ws.cell(row=current_row, column=12, value=record.get('sem2_cwts_f', 0))     # Column L (CWTS F)
        ws.cell(row=current_row, column=13, value=record.get('sem2_lts_m', 0))      # Column M (LTS M)
        ws.cell(row=current_row, column=14, value=record.get('sem2_lts_f', 0))      # Column N (LTS F)

        # SUMMER (Columns O to T)
        ws.cell(row=current_row, column=15, value=record.get('summer_rotc_m', 0))   # Column O
        ws.cell(row=current_row, column=16, value=record.get('summer_rotc_f', 0))   # Column P
        ws.cell(row=current_row, column=17, value=record.get('summer_cwts_m', 0))   # Column Q
        ws.cell(row=current_row, column=18, value=record.get('summer_cwts_f', 0))   # Column R
        ws.cell(row=current_row, column=19, value=record.get('summer_lts_m', 0))    # Column S
        ws.cell(row=current_row, column=20, value=record.get('summer_lts_f', 0))    # Column T

        # GRADUATES (Columns U to Z)
        ws.cell(row=current_row, column=21, value=record.get('grad_rotc_m', 0))     # Column U
        ws.cell(row=current_row, column=22, value=record.get('grad_rotc_f', 0))     # Column V
        ws.cell(row=current_row, column=23, value=record.get('grad_cwts_m', 0))     # Column W
        ws.cell(row=current_row, column=24, value=record.get('grad_cwts_f', 0))     # Column X
        ws.cell(row=current_row, column=25, value=record.get('grad_lts_m', 0))      # Column Y
        ws.cell(row=current_row, column=26, value=record.get('grad_lts_f', 0))      # Column Z

        # Apply borders and alignment to data row
        for c in range(1, 27):
            cell = ws.cell(row=current_row, column=c)
            cell.border = thin_border
            cell.font = data_font
            cell.alignment = align_left if c <= 2 else align_center

    # 4. Signatories Footer
    sig_row = start_row + len(database_records) + 2
    ws.cell(row=sig_row, column=1, value="Prepared by: NSTP Department Coordinator").font = Font(name="Arial", size=9, bold=True)
    ws.cell(row=sig_row, column=10, value="Verified by: Campus NSTP Director").font = Font(name="Arial", size=9, bold=True)
    ws.cell(row=sig_row, column=20, value="Approved by: Campus Administrator").font = Font(name="Arial", size=9, bold=True)
        
    # 5. I-save bilang BAGONG file para hindi masira ang original template
    output_filename = 'Ready_To_Print_OSDS-NSTP-Form-2-A.xlsx'
    wb.save(output_filename)
    
    print(f"Success! File is ready for download: {output_filename}")
    return output_filename


# --- EXAMPLE USAGE & TEST ---
if __name__ == '__main__':
    sample_db_records = [
        {
            'hei_name': 'CAVITE STATE UNIVERSITY - NAIC', 
            'classification': 'PUBLIC',
            'sem1_rotc_m': 45, 'sem1_rotc_f': 12,
            'sem1_cwts_m': 120, 'sem1_cwts_f': 150,
            'sem1_lts_m': 15, 'sem1_lts_f': 25,
            'sem2_rotc_m': 43, 'sem2_rotc_f': 12,
            'sem2_cwts_m': 118, 'sem2_cwts_f': 149,
            'sem2_lts_m': 15, 'sem2_lts_f': 25,
            'summer_rotc_m': 0, 'summer_rotc_f': 0,
            'summer_cwts_m': 0, 'summer_cwts_f': 0,
            'summer_lts_m': 0, 'summer_lts_f': 0,
            'grad_rotc_m': 43, 'grad_rotc_f': 12,
            'grad_cwts_m': 118, 'grad_cwts_f': 149,
            'grad_lts_m': 15, 'grad_lts_f': 25
        },
        {
            'hei_name': 'CVSU BACOOR', 
            'classification': 'PUBLIC',
            'sem1_rotc_m': 30, 'sem1_rotc_f': 10,
            'sem1_cwts_m': 90, 'sem1_cwts_f': 110,
            'sem1_lts_m': 0, 'sem1_lts_f': 0,
            'sem2_rotc_m': 30, 'sem2_rotc_f': 10,
            'sem2_cwts_m': 88, 'sem2_cwts_f': 108,
            'sem2_lts_m': 0, 'sem2_lts_f': 0,
            'summer_rotc_m': 0, 'summer_rotc_f': 0,
            'summer_cwts_m': 0, 'summer_cwts_f': 0,
            'summer_lts_m': 0, 'summer_lts_f': 0,
            'grad_rotc_m': 30, 'grad_rotc_f': 10,
            'grad_cwts_m': 88, 'grad_cwts_f': 108,
            'grad_lts_m': 0, 'grad_lts_f': 0
        }
    ]

    generate_ready_to_print_form_a(sample_db_records)
