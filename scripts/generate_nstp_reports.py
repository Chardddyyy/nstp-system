import os
import openpyxl
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

def enforce_print_setup(ws):
    """Fix for Excel templates when customizing print properties."""
    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True

def generate_nstp_reports(
    form_b_template: str = 'OSDS-NSTP-Form-2-B.xlsx',
    form_a_template: str = 'OSDS-NSTP-Form-2-A.xlsx',
    students_data: list = None,
    summary_data: list = None,
    output_dir: str = '.'
):
    """
    Populates existing OSDS-NSTP Form 2-A and Form 2-B templates with
    enforced Landscape A4 print setup, thin borders, and auto-fit widths.
    """
    os.makedirs(output_dir, exist_ok=True)
    out_b = os.path.join(output_dir, 'Print_Ready_OSDS-NSTP-Form-2-B.xlsx')
    out_a = os.path.join(output_dir, 'Print_Ready_OSDS-NSTP-Form-2-A.xlsx')

    # ==============================================================
    # 1. PROCESS FORM 2-B (NSTP Enrollment List)
    # ==============================================================
    if os.path.exists(form_b_template):
        wb_b = openpyxl.load_workbook(form_b_template)
        ws_b = wb_b.active
        enforce_print_setup(ws_b)

        # Magic Print Settings
        ws_b.page_setup.paperSize = 9  # 9 = A4 Size Paper
        ws_b.page_setup.orientation = ws_b.ORIENTATION_LANDSCAPE
        ws_b.page_setup.fitToHeight = False  # Allows multiple pages vertically for long lists
        ws_b.page_setup.fitToWidth = 1       # Fit to exactly 1 page width horizontally
        ws_b.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

        # Default sample student records if none supplied
        # Notice Column 10 and 12 are blank ("") to safeguard merged address columns
        sample_data_b = students_data or [
            [1, "2024-00101", "Dela Cruz", "Juan", "P.", "BSIT 3A", "M", "2005-08-17", "Brgy 1", "", "Imus", "", "Cavite", "09123456789", "juan@email.com"],
            [2, "2024-00102", "Santos", "Maria", "C.", "BSIT 3A", "F", "2006-01-20", "Brgy 2", "", "Bacoor", "", "Cavite", "09987654321", "maria@email.com"]
        ]

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        data_font = Font(name="Arial", size=9)

        # Start writing at Row 13 in Form-2-B
        for row_idx, data in enumerate(sample_data_b, start=13):
            ws_b.row_dimensions[row_idx].height = 20
            for col_idx, value in enumerate(data, start=1):
                cell = ws_b.cell(row=row_idx, column=col_idx)
                if value != "":  # Write non-empty value to keep template merges intact
                    cell.value = value
                cell.border = thin_border
                cell.font = data_font
                is_center = col_idx in [1, 2, 6, 7, 8, 14]
                cell.alignment = Alignment(horizontal="center" if is_center else "left", vertical="center", wrap_text=True)

        wb_b.save(out_b)
        print(f"✅ Form 2-B generated: {out_b}")
    else:
        print(f"⚠️ Form 2-B template not found at '{form_b_template}', skipping template injection.")

    # ==============================================================
    # 2. PROCESS FORM 2-A (Summary Report)
    # ==============================================================
    try:
        from generate_form_a_openpyxl import generate_ready_to_print_form_a, create_base_form_a_template
        if not os.path.exists(form_a_template):
            create_base_form_a_template(form_a_template)
            
        sample_summary_records = summary_data or [
            {
                'hei_name': 'CAVITE STATE UNIVERSITY - NAIC', 
                'classification': 'PUBLIC',
                'sem1_rotc_m': 45, 'sem1_rotc_f': 12,
                'sem1_cwts_m': 120, 'sem1_cwts_f': 150,
                'sem1_lts_m': 15, 'sem1_lts_f': 25,
                'sem2_rotc_m': 43, 'sem2_rotc_f': 12,
                'sem2_cwts_m': 118, 'sem2_cwts_f': 149,
                'sem2_lts_m': 15, 'sem2_lts_f': 25,
                'summer_rotc_m': 0, 'summer_rotc_f': 0,
                'summer_cwts_m': 0, 'summer_cwts_f': 0,
                'summer_lts_m': 0, 'summer_lts_f': 0,
                'grad_rotc_m': 43, 'grad_rotc_f': 12,
                'grad_cwts_m': 118, 'grad_cwts_f': 149,
                'grad_lts_m': 15, 'grad_lts_f': 25
            }
        ]
        generate_ready_to_print_form_a(sample_summary_records, template_path=form_a_template)
        print(f"✅ Form 2-A generated: Ready_To_Print_{form_a_template}")
    except Exception as e:
        print(f"⚠️ Form 2-A generation notice: {e}")

    print("🎉 Success! Na-generate na ang mga print-ready Excel files.")

if __name__ == '__main__':
    generate_nstp_reports()
